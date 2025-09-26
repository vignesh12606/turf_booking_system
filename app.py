import sqlite3
from flask import Flask, render_template, request, redirect, url_for, session, flash, g, Response
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import io
from fpdf import FPDF
import openpyxl


# --- APP SETUP ---
app = Flask(__name__)
# It's crucial to set a secret key for session management.
# In a production environment, use a more complex, securely stored key.
app.config['SECRET_KEY'] = 'a_very_secret_and_secure_key_for_turf_booking'
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database/turf_booking.db')

# --- DATABASE MANAGEMENT ---
def get_db():
    """Opens a new database connection if there is none yet for the current application context."""
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row  # This allows accessing columns by name
    return g.db

@app.teardown_appcontext
def close_db(exception):
    """Closes the database again at the end of the request."""
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    """Initializes the database using the schema.sql file."""
    with app.app_context():
        db = get_db()
        with app.open_resource('database/schema.sql', mode='r') as f:
            db.cursor().executescript(f.read())
        db.commit()

@app.cli.command('initdb')
def initdb_command():
    """Flask command to initialize the database."""
    init_db()
    print('Initialized the database.')

# --- MIDDLEWARE & HELPERS ---
@app.before_request
def before_request():
    """Pre-request logic to manage user sessions."""
    g.user = None
    if 'user_id' in session:
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        g.user = user

# --- AUTHENTICATION ROUTES ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        
        db = get_db()
        error = None

        if not username:
            error = 'Username is required.'
        elif not password:
            error = 'Password is required.'
        elif db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone() is not None:
            error = f"User {username} is already registered."

        if error is None:
            db.execute(
                'INSERT INTO users (username, email, password_hash) VALUES (?, ?, ?)',
                (username, email, generate_password_hash(password))
            )
            db.commit()
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        
        flash(error, 'error')

    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        error = None
        user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()

        if user is None:
            error = 'Incorrect username.'
        elif not check_password_hash(user['password_hash'], password):
            error = 'Incorrect password.'

        if error is None:
            session.clear()
            session['user_id'] = user['id']
            if user['is_admin']:
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('index'))

        flash(error, 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# --- USER ROUTES ---
@app.route('/')
def index():
    if g.user is None:
        return redirect(url_for('login'))
    
    db = get_db()
    turfs = db.execute('SELECT * FROM turfs').fetchall()
    return render_template('index.html', turfs=turfs)

@app.route('/turf/<int:turf_id>')
def turf_details(turf_id):
    if g.user is None:
        return redirect(url_for('login'))

    db = get_db()
    turf = db.execute('SELECT * FROM turfs WHERE id = ?', (turf_id,)).fetchone()
    
    # Generate next 7 days for booking
    today = datetime.now()
    dates = [(today + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
    
    # Generate time slots (e.g., from 9 AM to 9 PM)
    time_slots = [f"{h:02d}:00" for h in range(9, 22)] # 9 AM to 9 PM (21:00)

    return render_template('turf_details.html', turf=turf, dates=dates, time_slots=time_slots)

@app.route('/check_availability')
def check_availability():
    """API endpoint to check if a slot is booked."""
    turf_id = request.args.get('turf_id')
    date = request.args.get('date')
    time = request.args.get('time')
    booking_datetime_str = f"{date} {time}"
    
    db = get_db()
    booking = db.execute(
        'SELECT id FROM bookings WHERE turf_id = ? AND booking_time = ? AND status = "Confirmed"',
        (turf_id, booking_datetime_str)
    ).fetchone()
    
    return {'available': booking is None}

# --- BOOKING WORKFLOW ---
@app.route('/book/confirm', methods=['POST'])
def confirm_booking():
    """Step 1: Calculate price and show confirmation page."""
    if g.user is None:
        return redirect(url_for('login'))

    turf_id = request.form.get('turf_id')
    date = request.form.get('date')
    time = request.form.get('time')
    redeem_points = 'redeem_points' in request.form

    if not all([turf_id, date, time]):
        flash('Missing booking information. Please select a date and time.', 'error')
        return redirect(request.referrer or url_for('index'))
    
    db = get_db()
    turf = db.execute('SELECT * FROM turfs WHERE id = ?', (turf_id,)).fetchone()
    if turf is None:
        flash('Selected turf not found.', 'error')
        return redirect(url_for('index'))

    amount = turf['price_per_hour']
    points_to_redeem = 0
    discount = 0.0

    if redeem_points and g.user['loyalty_points'] >= 50:
        points_to_redeem = 50
        discount = amount * 0.25
    
    final_amount = amount - discount

    booking_details = {
        "date": date,
        "time": time,
        "final_amount": final_amount,
        "discount": discount,
        "points_to_redeem": points_to_redeem
    }

    return render_template('payment_confirmation.html', turf=turf, booking_details=booking_details)

@app.route('/book/execute', methods=['POST'])
def execute_booking():
    """Step 2: Process the payment and save the booking."""
    if g.user is None:
        return redirect(url_for('login'))

    db = get_db()
    turf_id = request.form.get('turf_id')
    date = request.form.get('date')
    time = request.form.get('time')
    final_amount = float(request.form.get('final_amount'))
    points_redeemed = int(request.form.get('points_redeemed'))

    booking_datetime = f"{date} {time}"

    # Final check: Ensure the slot has not been booked by someone else
    # while the user was on the confirmation page. This prevents duplicate bookings.
    existing_booking = db.execute(
        'SELECT id FROM bookings WHERE turf_id = ? AND booking_time = ? AND status = "Confirmed"',
        (turf_id, booking_datetime)
    ).fetchone()

    if existing_booking:
        flash('Sorry, this slot was just booked by another user. Please select a different time.', 'error')
        return redirect(url_for('turf_details', turf_id=turf_id))

    turf = db.execute('SELECT * FROM turfs WHERE id = ?', (turf_id,)).fetchone()

    # Update user's points
    current_points = g.user['loyalty_points']
    new_points = (current_points - points_redeemed) + 10
    
    # Insert booking
    db.execute(
        'INSERT INTO bookings (user_id, turf_id, booking_time, amount_paid, points_redeemed) VALUES (?, ?, ?, ?, ?)',
        (g.user['id'], turf_id, booking_datetime, final_amount, points_redeemed)
    )
    
    # Update user's points balance
    db.execute('UPDATE users SET loyalty_points = ? WHERE id = ?', (new_points, g.user['id']))
    
    db.commit()

    flash('Booking successful! You have earned 10 loyalty points.', 'success')
    return render_template('booking_confirmation.html', turf=turf, date=date, time=time, amount=final_amount, discount=(turf['price_per_hour'] - final_amount))

@app.route('/dashboard')
def dashboard():
    if g.user is None:
        return redirect(url_for('login'))

    db = get_db()
    bookings = db.execute(
        '''
        SELECT b.id, t.name, t.location, b.booking_time, b.status, b.amount_paid
        FROM bookings b
        JOIN turfs t ON b.turf_id = t.id
        WHERE b.user_id = ?
        ORDER BY b.booking_time DESC
        ''', (g.user['id'],)
    ).fetchall()

    return render_template('dashboard.html', bookings=bookings)

@app.route('/cancel_booking/<int:booking_id>', methods=['POST'])
def cancel_booking(booking_id):
    if g.user is None:
        return redirect(url_for('login'))

    db = get_db()
    booking = db.execute('SELECT * FROM bookings WHERE id = ? AND user_id = ?', (booking_id, g.user['id'])).fetchone()
    
    if booking:
        db.execute('UPDATE bookings SET status = "Cancelled" WHERE id = ?', (booking_id,))
        
        # Logic to return/remove points
        points_change = 0
        if booking['points_redeemed'] > 0:
            points_change += booking['points_redeemed'] # Return redeemed points
        points_change -= 10 # Remove awarded points
        
        if points_change != 0:
             db.execute('UPDATE users SET loyalty_points = loyalty_points + ? WHERE id = ?', (points_change, g.user['id']))

        db.commit()
        flash('Booking has been cancelled.', 'success')
    else:
        flash('Booking not found or you do not have permission to cancel it.', 'error')
        
    return redirect(url_for('dashboard'))

# --- ADMIN ROUTES ---
@app.route('/admin')
def admin_dashboard():
    if not g.user or not g.user['is_admin']:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('index'))
    
    db = get_db()
    all_bookings = db.execute(
        '''
        SELECT b.id, u.username, t.name, b.booking_time, b.status, b.amount_paid
        FROM bookings b
        JOIN users u ON b.user_id = u.id
        JOIN turfs t ON b.turf_id = t.id
        ORDER BY b.booking_time DESC
        LIMIT 10
        '''
    ).fetchall()
    
    turfs = db.execute('SELECT * FROM turfs ORDER BY name').fetchall()
    return render_template('admin_dashboard.html', bookings=all_bookings, turfs=turfs)

@app.route('/admin/turf/add', methods=['POST'])
def add_turf():
    if not g.user or not g.user['is_admin']:
        return redirect(url_for('index'))
    
    name = request.form['name']
    location = request.form['location']
    description = request.form['description']
    price = request.form['price']
    image_url = request.form['image_url']
    
    db = get_db()
    db.execute(
        'INSERT INTO turfs (name, location, description, price_per_hour, image_url) VALUES (?, ?, ?, ?, ?)',
        (name, location, description, price, image_url)
    )
    db.commit()
    flash(f'Turf "{name}" added successfully.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/turf/remove/<int:turf_id>', methods=['POST'])
def remove_turf(turf_id):
    if not g.user or not g.user['is_admin']:
        return redirect(url_for('index'))
    
    db = get_db()
    # First, delete bookings associated with the turf to avoid foreign key errors
    db.execute('DELETE FROM bookings WHERE turf_id = ?', (turf_id,))
    # Then, delete the turf itself
    db.execute('DELETE FROM turfs WHERE id = ?', (turf_id,))
    db.commit()
    flash(f'Turf and all its associated bookings have been removed.', 'success')
    return redirect(url_for('admin_dashboard'))

# --- NEW REPORTING ROUTES ---
class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Turf Booking Report', 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

@app.route('/admin/report/pdf')
def download_pdf_report():
    if not g.user or not g.user['is_admin']:
        return redirect(url_for('index'))
    
    db = get_db()
    bookings = db.execute(
        '''
        SELECT u.username, t.name, b.booking_time, b.status, b.amount_paid
        FROM bookings b
        JOIN users u ON b.user_id = u.id
        JOIN turfs t ON b.turf_id = t.id
        ORDER BY b.booking_time DESC
        '''
    ).fetchall()
    
    pdf = PDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 10)
    
    # Table Header
    pdf.cell(40, 10, 'User', 1)
    pdf.cell(50, 10, 'Turf Name', 1)
    pdf.cell(50, 10, 'Booking Time', 1)
    pdf.cell(20, 10, 'Status', 1)
    pdf.cell(30, 10, 'Amount Paid', 1)
    pdf.ln()
    
    # Table Rows
    pdf.set_font('Arial', '', 10)
    for booking in bookings:
        pdf.cell(40, 10, str(booking['username']), 1)
        pdf.cell(50, 10, str(booking['name']), 1)
        pdf.cell(50, 10, str(booking['booking_time']), 1)
        pdf.cell(20, 10, str(booking['status']), 1)
        pdf.cell(30, 10, f"Rs.{booking['amount_paid']:.2f}", 1)
        pdf.ln()
        
    return Response(pdf.output(dest='S'),
                    mimetype='application/pdf',
                    headers={'Content-Disposition':'attachment;filename=booking_report.pdf'})

@app.route('/admin/report/excel')
def download_excel_report():
    if not g.user or not g.user['is_admin']:
        return redirect(url_for('index'))
        
    db = get_db()
    bookings = db.execute(
        '''
        SELECT u.username, t.name, b.booking_time, b.status, b.amount_paid
        FROM bookings b
        JOIN users u ON b.user_id = u.id
        JOIN turfs t ON b.turf_id = t.id
        ORDER BY b.booking_time DESC
        '''
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings Report"
    
    # Header
    ws.append(['Username', 'Turf Name', 'Booking Time', 'Status', 'Amount Paid (Rs)'])
    
    # Data
    for booking in bookings:
        ws.append([
            booking['username'],
            booking['name'],
            booking['booking_time'],
            booking['status'],
            booking['amount_paid']
        ])
    
    # Save to a memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(buffer,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition':'attachment;filename=booking_report.xlsx'})


# --- MAIN EXECUTION ---
if __name__ == '__main__':
    # Check if the database directory exists, if not create it
    if not os.path.exists(os.path.dirname(DATABASE)):
        os.makedirs(os.path.dirname(DATABASE))
        
    if not os.path.exists(DATABASE):
        print("Database not found. Initializing...")
        init_db()
    app.run(debug=True)
